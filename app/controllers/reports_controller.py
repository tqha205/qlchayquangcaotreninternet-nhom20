"""
Module 5 — Báo cáo & Xuất file
Blueprint: reports_bp (prefix /admin/reports)
"""
from flask import (Blueprint, request, jsonify, session,
                   render_template, redirect, url_for, send_file, Response)
from app.models import DBModel
from app.controllers.admin_controller import require_role_page, require_role_api
from datetime import datetime, timedelta
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import reportlab.lib.colors as rlc
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

reports_bp = Blueprint('reports', __name__)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _date_range_params():
    """Trả về (from_date, to_date) dạng string YYYY-MM-DD."""
    today    = datetime.today()
    from_str = request.args.get('from') or (today - timedelta(days=29)).strftime('%Y-%m-%d')
    to_str   = request.args.get('to')   or today.strftime('%Y-%m-%d')
    return from_str, to_str


def _get_summary(from_date, to_date, role, user_id, cust_id):
    """Trả về dict tổng hợp và list chi tiết. Anchor dựa trên dữ liệu thực tế (Daily Reports)."""
    where_role = ""
    join_role = ""
    params = [from_date, to_date]

    if role == 'client' and cust_id:
        where_role = "AND c.customer_id = %s"
        params.append(cust_id)
    elif role == 'marketer':
        where_role = "AND cu.marketer_id = %s"
        join_role = "JOIN customers cu ON c.customer_id = cu.id"
        params.append(user_id)

    # 1. Tổng hợp: Lấy các chiến dịch CÓ DỮ LIỆU trong kỳ
    sql_summary = f"""
        SELECT
            COUNT(DISTINCT c.id)                  AS total_campaigns,
            COALESCE(SUM(dr.daily_spent), 0)     AS total_spent
        FROM daily_reports dr
        JOIN campaigns c ON dr.campaign_id = c.id
        {join_role}
        WHERE dr.report_date >= %s AND dr.report_date <= %s
          AND c.is_deleted = 0
          {where_role}
    """
    summary = DBModel.fetch_one(sql_summary, tuple(params))

    # Tính tổng budget riêng để tránh bị nhân đôi do JOIN với daily_reports
    sql_budget = f"""
        SELECT COALESCE(SUM(budget), 0) AS total_budget
        FROM campaigns c
        {join_role}
        WHERE c.id IN (
            SELECT DISTINCT campaign_id 
            FROM daily_reports 
            WHERE report_date >= %s AND report_date <= %s
        )
        AND c.is_deleted = 0
        {where_role}
    """
    budget_data = DBModel.fetch_one(sql_budget, tuple(params))

    # 2. Breakdown theo platform
    sql_platform = f"""
        SELECT c.platform, COUNT(DISTINCT c.id) AS cnt,
               COALESCE(SUM(dr.daily_spent), 0) AS spent
        FROM daily_reports dr
        JOIN campaigns c ON dr.campaign_id = c.id
        {join_role}
        WHERE dr.report_date >= %s AND dr.report_date <= %s
          AND c.is_deleted = 0
          {where_role}
        GROUP BY c.platform
        ORDER BY spent DESC
    """
    platforms = DBModel.fetch_all(sql_platform, tuple(params))

    # 3. Chi tiết từng chiến dịch
    sql_detail = f"""
        SELECT c.id, c.name, cu.name AS customer_name, c.platform,
               c.budget, COALESCE(SUM(dr.daily_spent), 0) as spent, c.status, c.start_date, c.end_date,
               ROUND(COALESCE(SUM(dr.daily_spent), 0)/NULLIF(c.budget,0)*100, 1) AS spent_pct
        FROM daily_reports dr
        JOIN campaigns c ON dr.campaign_id = c.id
        LEFT JOIN customers cu ON c.customer_id = cu.id
        WHERE dr.report_date >= %s AND dr.report_date <= %s
          AND c.is_deleted = 0
          {where_role}
        GROUP BY c.id
        ORDER BY spent DESC, c.id DESC
    """
    details = DBModel.fetch_all(sql_detail, tuple(params))

    def safe_float(v):
        try: return float(v or 0)
        except: return 0.0

    total_spent = safe_float(summary['total_spent'])
    total_budget = safe_float(budget_data['total_budget'])

    return {
        'summary': {
            'total_campaigns': int(summary['total_campaigns'] or 0),
            'total_budget':    total_budget,
            'total_spent':     total_spent,
            'avg_ratio':       round(total_spent / total_budget * 100, 1) if total_budget > 0 else 0,
        },
        'platforms': [
            {
                'platform': p['platform'],
                'cnt':      int(p['cnt']),
                'budget':   0, # Chúng ta không tính budget lẻ cho platform ở đây để tránh phức tạp, table sẽ chỉ hiện spent
                'spent':    safe_float(p['spent']),
            }
            for p in platforms
        ],
        'details': [
            {
                'id':            d['id'],
                'name':          d['name'],
                'customer_name': d['customer_name'] or '—',
                'platform':      d['platform'],
                'budget':        safe_float(d['budget']),
                'spent':         safe_float(d['spent']),
                'status':        d['status'],
                'start_date':    str(d['start_date']) if d['start_date'] else '—',
                'end_date':      str(d['end_date'])   if d['end_date']   else '—',
                'spent_pct':     safe_float(d['spent_pct']),
            }
            for d in details
        ],
    }


# ─────────────────────────────────────────────────────────────────────────────
# Page Route
# ─────────────────────────────────────────────────────────────────────────────

@reports_bp.route('/')
@require_role_page(['admin', 'marketer', 'client'])
def reports_page():
    return render_template('admin/reports.html')


# ─────────────────────────────────────────────────────────────────────────────
# API — Summary
# ─────────────────────────────────────────────────────────────────────────────

@reports_bp.route('/api/summary')
@require_role_api(['admin', 'marketer', 'client'])
def reports_summary():
    from_date, to_date = _date_range_params()
    role    = session['role']
    user_id = session.get('user_id')
    cust_id = session.get('customer_id')
    data    = _get_summary(from_date, to_date, role, user_id, cust_id)
    return jsonify({'success': True, 'from': from_date, 'to': to_date, **data})

@reports_bp.route('/api/chart/spending-trend')
@require_role_api(['admin', 'marketer', 'client'])
def chart_spending_trend():
    from_date, to_date = _date_range_params()
    role    = session['role']
    cust_id = session.get('customer_id')

    where_extra = "AND c.customer_id = %s" if role == 'client' and cust_id else ""
    where_marketer = "AND cu.marketer_id = %s" if role == 'marketer' else ""
    join_marketer = "JOIN customers cu ON c.customer_id = cu.id" if role == 'marketer' else ""
    
    params = [from_date, to_date]
    if role == 'client' and cust_id: params.append(cust_id)
    elif role == 'marketer': params.append(session.get('user_id'))

    sql = f"""
        SELECT dr.report_date as report_date, SUM(dr.daily_spent) AS total_spent
        FROM daily_reports dr
        JOIN campaigns c ON dr.campaign_id = c.id
        {join_marketer}
        WHERE dr.report_date >= %s AND dr.report_date <= %s
          AND c.is_deleted = 0
          {where_extra}
          {where_marketer}
        GROUP BY dr.report_date
        ORDER BY dr.report_date ASC
    """
    rows = DBModel.fetch_all(sql, tuple(params))
    
    labels = [r['report_date'].strftime('%d/%m') if hasattr(r['report_date'], 'strftime') else str(r['report_date']) for r in rows]
    data   = [float(r['total_spent'] or 0) for r in rows]
    return jsonify({'success': True, 'labels': labels, 'data': data})


@reports_bp.route('/api/agency/profit-report')
@require_role_api(['admin'])
def agency_profit_report():
    """
    SQL Tối ưu: Lấy báo cáo doanh thu & lợi nhuận ròng theo tháng.
    Doanh thu = Tổng chi tiêu của khách * 10% (phí dịch vụ).
    Lợi nhuận = Doanh thu - (Chi phí vận hành giả định 2%).
    """
    sql = """
        SELECT 
            DATE_FORMAT(dr.report_date, '%Y-%m') as month,
            SUM(dr.daily_spent) as total_client_spend,
            SUM(dr.daily_spent * 0.1) as agency_revenue, -- 10% Service Fee
            SUM(dr.daily_spent * 0.08) as net_profit     -- Trừ 2% vận hành -> còn 8% thực thu
        FROM daily_reports dr
        JOIN campaigns c ON dr.campaign_id = c.id
        WHERE dr.report_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
          AND c.is_deleted = 0
        GROUP BY month
        ORDER BY month DESC
    """
    data = DBModel.fetch_all(sql)
    return jsonify({'success': True, 'data': data})


# ─────────────────────────────────────────────────────────────────────────────
# API — Export Excel
# ─────────────────────────────────────────────────────────────────────────────

@reports_bp.route('/api/export/excel')
@require_role_api(['admin', 'marketer', 'client'])
def export_excel():
    if not OPENPYXL_OK:
        return jsonify({'success': False, 'message': 'Thư viện openpyxl chưa được cài. Chạy: pip install openpyxl'}), 500

    from_date, to_date = _date_range_params()
    role    = session['role']
    cust_id = session.get('customer_id')
    data    = _get_summary(from_date, to_date, role, cust_id)

    wb = openpyxl.Workbook()

    # ── Style helpers ──────────────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill    = PatternFill('solid', fgColor='F8FAFC')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),  bottom=Side(style='thin', color='E2E8F0')
    )

    def set_header(ws, row, cols):
        for col, title in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col, value=title)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin_border

    def fmt_vnd(val):
        return f"{val:,.0f} VNĐ"

    # ── Sheet 1: Tổng hợp ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Tổng hợp'
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20

    ws1['A1'] = 'BÁO CÁO TỔNG HỢP CHIẾN DỊCH QUẢNG CÁO'
    ws1['A1'].font      = Font(bold=True, size=14, color='1E293B')
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A1:B1')

    ws1['A2'] = f'Kỳ báo cáo: {from_date} → {to_date}'
    ws1['A2'].font      = Font(italic=True, size=10, color='64748B')
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:B2')

    kpi_rows = [
        ('Tổng số chiến dịch',    data['summary']['total_campaigns']),
        ('Tổng ngân sách',        fmt_vnd(data['summary']['total_budget'])),
        ('Tổng chi tiêu',         fmt_vnd(data['summary']['total_spent'])),
        ('Tỷ lệ sử dụng TB',      f"{data['summary']['avg_ratio']}%"),
    ]
    for r, (label, value) in enumerate(kpi_rows, 4):
        ws1[f'A{r}'] = label
        ws1[f'B{r}'] = value
        ws1[f'A{r}'].font = Font(bold=True)
        ws1[f'B{r}'].alignment = Alignment(horizontal='right')

    ws1['A9'] = 'Phân bổ theo nền tảng'
    ws1['A9'].font = Font(bold=True, size=12)
    set_header(ws1, 10, ['Nền tảng', 'Số CĐ', 'Ngân sách', 'Chi tiêu'])
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20

    for r, p in enumerate(data['platforms'], 11):
        row_data = [p['platform'], p['cnt'], fmt_vnd(p['budget']), fmt_vnd(p['spent'])]
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center' if c == 2 else 'left')
            if r % 2 == 0:
                cell.fill = alt_fill

    # ── Sheet 2: Chi tiết chiến dịch ──────────────────────────────────────
    ws2 = wb.create_sheet('Chi tiết chiến dịch')
    cols  = ['STT', 'Tên chiến dịch', 'Khách hàng', 'Nền tảng',
             'Ngân sách (VNĐ)', 'Đã chi (VNĐ)', 'Tỷ lệ (%)', 'Trạng thái', 'Bắt đầu', 'Kết thúc']
    widths = [6, 30, 22, 15, 18, 18, 12, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    set_header(ws2, 1, cols)

    for r, d in enumerate(data['details'], 2):
        row_data = [
            r - 1, d['name'], d['customer_name'], d['platform'],
            d['budget'], d['spent'], d['spent_pct'], d['status'],
            d['start_date'], d['end_date'],
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center' if c in (1, 7) else 'left')
            if r % 2 == 0:
                cell.fill = alt_fill

    # ── Output ──────────────────────────────────────────────────────────────
    buf      = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"bao-cao-{from_date}--{to_date}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ─────────────────────────────────────────────────────────────────────────────
# API — Export PDF
# ─────────────────────────────────────────────────────────────────────────────

@reports_bp.route('/api/export/pdf')
@require_role_api(['admin', 'marketer', 'client'])
def export_pdf():
    if not REPORTLAB_OK:
        return jsonify({'success': False, 'message': 'Thư viện reportlab chưa được cài. Chạy: pip install reportlab'}), 500

    from_date, to_date = _date_range_params()
    role    = session['role']
    cust_id = session.get('customer_id')
    data    = _get_summary(from_date, to_date, role, cust_id)

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                             leftMargin=1.5*cm, rightMargin=1.5*cm)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 fontSize=16, textColor=rlc.HexColor('#1E293B'),
                                 spaceAfter=4, alignment=TA_CENTER)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=9, textColor=rlc.HexColor('#64748B'),
                                 spaceAfter=16, alignment=TA_CENTER)
    section_style = ParagraphStyle('section', parent=styles['Heading2'],
                                   fontSize=11, textColor=rlc.HexColor('#4F46E5'),
                                   spaceBefore=12, spaceAfter=6)

    # ── Header ────────────────────────────────────────────────────────────
    story.append(Paragraph('BÁO CÁO CHIẾN DỊCH QUẢNG CÁO', title_style))
    story.append(Paragraph(f'Kỳ báo cáo: {from_date} → {to_date} | Xuất lúc: {datetime.now().strftime("%d/%m/%Y %H:%M")}', sub_style))

    # ── KPI Table ─────────────────────────────────────────────────────────
    s = data['summary']
    kpi_data = [
        ['Chỉ số', 'Giá trị'],
        ['Tổng số chiến dịch',     str(s['total_campaigns'])],
        ['Tổng ngân sách',         f"{s['total_budget']:,.0f} VNĐ"],
        ['Tổng chi tiêu',          f"{s['total_spent']:,.0f} VNĐ"],
        ['Tỷ lệ sử dụng ngân sách TB', f"{s['avg_ratio']}%"],
    ]
    story.append(Paragraph('I. Tổng quan', section_style))
    kpi_table = Table(kpi_data, colWidths=[9*cm, 9*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  rlc.HexColor('#4F46E5')),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  rlc.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rlc.white, rlc.HexColor('#F8FAFC')]),
        ('GRID',        (0, 0), (-1, -1), 0.5, rlc.HexColor('#E2E8F0')),
        ('TOPPADDING',  (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',(0,0), (-1, -1), 6),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.4*cm))

    # ── Platform Breakdown ────────────────────────────────────────────────
    story.append(Paragraph('II. Phân bổ theo nền tảng', section_style))
    plat_data = [['Nền tảng', 'Số chiến dịch', 'Ngân sách (VNĐ)', 'Chi tiêu (VNĐ)']]
    for p in data['platforms']:
        plat_data.append([p['platform'], str(p['cnt']),
                          f"{p['budget']:,.0f}", f"{p['spent']:,.0f}"])
    if len(plat_data) == 1:
        plat_data.append(['—', '—', '—', '—'])
    plat_table = Table(plat_data, colWidths=[4.5*cm, 3.5*cm, 5.5*cm, 5.5*cm])
    plat_table.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  rlc.HexColor('#4F46E5')),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  rlc.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 8),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rlc.white, rlc.HexColor('#F8FAFC')]),
        ('GRID',        (0, 0), (-1, -1), 0.5, rlc.HexColor('#E2E8F0')),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',(0,0), (-1, -1), 5),
    ]))
    story.append(plat_table)
    story.append(Spacer(1, 0.4*cm))

    # ── Detail Table ──────────────────────────────────────────────────────
    story.append(Paragraph('III. Chi tiết từng chiến dịch', section_style))
    det_headers = ['STT', 'Tên chiến dịch', 'KH', 'Nền tảng', 'Ngân sách', 'Chi tiêu', 'Tỷ lệ', 'Trạng thái']
    det_data = [det_headers]
    for i, d in enumerate(data['details'], 1):
        det_data.append([
            str(i), d['name'][:20], d['customer_name'][:14], d['platform'],
            f"{d['budget']:,.0f}", f"{d['spent']:,.0f}",
            f"{d['spent_pct']:.1f}%", d['status'],
        ])
    if len(det_data) == 1:
        det_data.append(['—'] * 8)

    col_widths = [1*cm, 4.5*cm, 3*cm, 2.5*cm, 3*cm, 3*cm, 1.5*cm, 2.5*cm]
    det_table  = Table(det_data, colWidths=col_widths)
    det_table.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  rlc.HexColor('#4F46E5')),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  rlc.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 7),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rlc.white, rlc.HexColor('#F8FAFC')]),
        ('GRID',        (0, 0), (-1, -1), 0.5, rlc.HexColor('#E2E8F0')),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0,0), (-1, -1), 4),
        ('WORDWRAP',    (1, 1), (1, -1),  True),
    ]))
    story.append(det_table)

    doc.build(story)
    buf.seek(0)
    filename = f"bao-cao-{from_date}--{to_date}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)
